from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import csv
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

Q = Decimal("0.01")


def money(value):
    return float(Decimal(str(value or 0)).quantize(Q, rounding=ROUND_HALF_UP))


def number(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        raise ValueError("Numeric fields must contain valid numbers")


def parse_time(value):
    if not value:
        return None

    value = str(value).strip()

    # User can enter:
    # 13:00 PM
    # 1:00 PM
    # 01:00 pm
    # 13:00
    # 1:00
    formats = (
        "%H:%M %p",
        "%H:%M:%S %p",
        "%I:%M %p",
        "%I:%M:%S %p",
        "%H:%M",
        "%H:%M:%S",
        "%I:%M",
        "%I:%M:%S",
    )

    for fmt in formats:
        try:
            return datetime.strptime(value.upper(), fmt)
        except ValueError:
            pass

    raise ValueError(
        f"Invalid time format: {value}. "
        "Please enter time like 13:00 PM or 1:00 PM."
    )


def calculate_trip(raw):
    start_km = number(raw.get("start_km"))
    end_km = number(raw.get("end_km"))
    if end_km < start_km:
        raise ValueError("End KM cannot be less than Start KM")

    start_time = parse_time(raw.get("start_time"))
    end_time = parse_time(raw.get("end_time"))
    total_hours = 0.0
    if start_time and end_time:
        delta = end_time - start_time
        if delta.total_seconds() < 0:
            delta = delta.replace(days=1)
        total_hours = delta.total_seconds() / 3600

    total_km = end_km - start_km
    slab_hours = max(0.0, number(raw.get("slab_hours")))
    slab_km = max(0.0, number(raw.get("slab_km")))
    slab_rate = max(0.0, number(raw.get("slab_rate")))
    extra_hour_rate = max(0.0, number(raw.get("extra_hour_rate")))
    extra_km_rate = max(0.0, number(raw.get("extra_km_rate")))
    driver_bata = max(0.0, number(raw.get("driver_bata")))
    parking = max(0.0, number(raw.get("parking")))
    toll = max(0.0, number(raw.get("toll")))
    other_charges = max(0.0, number(raw.get("other_charges")))

    extra_hours = max(0.0, total_hours - slab_hours) if slab_hours > 0 else 0.0
    extra_km = max(0.0, total_km - slab_km) if slab_km > 0 else 0.0
    extra_hour_amount = extra_hours * extra_hour_rate
    extra_km_amount = extra_km * extra_km_rate
    base_amount = slab_rate
    trip_total = base_amount + extra_hour_amount + extra_km_amount + driver_bata + parking + toll + other_charges

    return {
        "ds_no": str(raw.get("ds_no") or "").strip(),
        "trip_date": raw.get("trip_date"),
        "vehicle_type": str(raw.get("vehicle_type") or "").strip(),
        "vehicle_number": str(raw.get("vehicle_number") or "").strip(),
        "start_time": raw.get("start_time") or "",
        "end_time": raw.get("end_time") or "",
        "start_km": money(start_km),
        "end_km": money(end_km),
        "total_hours": money(total_hours),
        "total_km": money(total_km),
        "slab_hours": money(slab_hours),
        "slab_km": money(slab_km),
        "slab_rate": money(slab_rate),
        "extra_hour_rate": money(extra_hour_rate),
        "extra_km_rate": money(extra_km_rate),
        "extra_hours": money(extra_hours),
        "extra_km": money(extra_km),
        "extra_hour_amount": money(extra_hour_amount),
        "extra_km_amount": money(extra_km_amount),
        "base_amount": money(base_amount),
        "driver_bata": money(driver_bata),
        "parking": money(parking),
        "toll": money(toll),
        "other_charges": money(other_charges),
        "trip_total": money(trip_total),
        "notes": str(raw.get("notes") or "").strip(),
    }


def calculate_invoice(invoice_data, trip_data):
    trips = [calculate_trip(t) for t in trip_data]
    subtotal = sum(t["trip_total"] for t in trips)
    cgst_rate = max(0.0, number(invoice_data.get("cgst_rate")))
    sgst_rate = max(0.0, number(invoice_data.get("sgst_rate")))
    igst_rate = max(0.0, number(invoice_data.get("igst_rate")))
    cgst = subtotal * cgst_rate / 100
    sgst = subtotal * sgst_rate / 100
    igst = subtotal * igst_rate / 100
    exact_total = subtotal + cgst + sgst + igst
    rounded_total = float(Decimal(str(exact_total)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    round_off = rounded_total - exact_total
    return trips, {
        "cgst_rate": cgst_rate,
        "sgst_rate": sgst_rate,
        "igst_rate": igst_rate,
        "subtotal": money(subtotal),
        "cgst": money(cgst),
        "sgst": money(sgst),
        "igst": money(igst),
        "round_off": money(round_off),
        "grand_total": money(rounded_total),
    }


def amount_to_words_indian(amount):
    n = int(round(float(amount)))
    if n == 0:
        return "Rupees Zero Only"

    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def two(x):
        if x < 20:
            return ones[x]
        return tens[x // 10] + ((" " + ones[x % 10]) if x % 10 else "")

    def three(x):
        if x < 100:
            return two(x)
        return ones[x // 100] + " Hundred" + ((" " + two(x % 100)) if x % 100 else "")

    parts = []
    crore = n // 10000000
    n %= 10000000
    lakh = n // 100000
    n %= 100000
    thousand = n // 1000
    n %= 1000
    hundred = n
    if crore:
        parts.append(three(crore) + " Crore")
    if lakh:
        parts.append(three(lakh) + " Lakh")
    if thousand:
        parts.append(three(thousand) + " Thousand")
    if hundred:
        parts.append(three(hundred))
    return "Rupees " + " ".join(parts) + " Only"


def _p(text, style):
    return Paragraph(str(text if text is not None else ""), style)


# def build_invoice_pdf(invoice):
    buffer = BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        rightMargin=9 * mm,
        leftMargin=9 * mm,
        topMargin=7 * mm,
        bottomMargin=7 * mm,
        title=f"PVR Invoice {invoice.invoice_number}",
        author="PVR Tours & Travels",
    )

    styles = getSampleStyleSheet()
    company = ParagraphStyle("company", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=17, leading=18, alignment=TA_CENTER, spaceAfter=1)
    small_center = ParagraphStyle("small_center", parent=styles["Normal"], fontSize=7, leading=8, alignment=TA_CENTER)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=6.2, leading=7)
    tiny = ParagraphStyle("tiny", parent=styles["Normal"], fontSize=5.3, leading=6)
    tiny_bold = ParagraphStyle("tiny_bold", parent=tiny, fontName="Helvetica-Bold")
    section = ParagraphStyle("section", parent=styles["Normal"], fontSize=8, leading=9, fontName="Helvetica-Bold")
    right = ParagraphStyle("right", parent=tiny, alignment=TA_RIGHT)
    total_style = ParagraphStyle("total", parent=styles["Normal"], fontSize=7, leading=8, fontName="Helvetica-Bold", alignment=TA_RIGHT)

    story = []
    story.append(_p("P.V.R. TOURS AND TRAVELS", company))
    story.append(_p("H. No. 4-1-756, Tuljaguda, Troop Bazar, Hyderabad, Telangana State - 500 001.  Ph: 9030588882 / 9963578399", small_center))
    story.append(Spacer(1, 2 * mm))
    story.append(Table([[_p("TAXABLE INVOICE", section)]], colWidths=[doc.width], style=[
        ("BOX", (0, 0), (-1, -1), 0.7, colors.black), ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2)
    ]))

    left = [
        ["Billed To", invoice.customer_name or ""],
        ["Address", invoice.customer_address or ""],
        ["GSTIN", invoice.customer_gstin or ""],
        ["Booked By", invoice.booked_by or ""],
        ["Used By", invoice.used_by or ""],
        ["Reference / PO", invoice.reference_number or ""],
    ]
    right_meta = [
        ["Invoice Date", invoice.invoice_date.strftime("%d-%m-%Y") if invoice.invoice_date else ""],
        ["Invoice No", invoice.invoice_number],
        # ["Series", invoice.invoice_series],
        # ["Serial No", invoice.invoice_serial_number],meenakshi
        ["GSTIN", "36AYPPR7981L1Z8"],
    ]
    info_data = [
        [Table([[_p(k, tiny_bold), _p(v, tiny)] for k, v in left], colWidths=[27 * mm, 115 * mm], style=[("VALIGN", (0,0),(-1,-1),"TOP"), ("BOX",(0,0),(-1,-1),0.3,colors.grey), ("INNERGRID",(0,0),(-1,-1),0.2,colors.lightgrey), ("LEFTPADDING",(0,0),(-1,-1),2), ("RIGHTPADDING",(0,0),(-1,-1),2), ("TOPPADDING",(0,0),(-1,-1),1), ("BOTTOMPADDING",(0,0),(-1,-1),1)]),
         Table([[_p(k, tiny_bold), _p(v, tiny)] for k, v in right_meta], colWidths=[28 * mm, 54 * mm], style=[("VALIGN", (0,0),(-1,-1),"TOP"), ("BOX",(0,0),(-1,-1),0.3,colors.grey), ("INNERGRID",(0,0),(-1,-1),0.2,colors.lightgrey), ("LEFTPADDING",(0,0),(-1,-1),2), ("RIGHTPADDING",(0,0),(-1,-1),2), ("TOPPADDING",(0,0),(-1,-1),1), ("BOTTOMPADDING",(0,0),(-1,-1),1)])]
    ]
    story.append(Table(info_data, colWidths=[doc.width * 0.73, doc.width * 0.27], style=[("VALIGN", (0,0),(-1,-1),"TOP"), ("LEFTPADDING",(0,0),(-1,-1),0), ("RIGHTPADDING",(0,0),(-1,-1),0)]))
    story.append(Spacer(1, 2 * mm))

    headers = [
        "DS No", "Date", "Car Type", "Car No", "Slab Hrs", "Slab Kms", "Slab Rate", "Ex. Hrs Rate", "Ex. Kms Rate",
        "Start Time", "End Time", "Start KMS", "End KMS", "Driver Bata", "Parking & Toll", "Total Hrs", "Ex. Hrs", "Ex. Hrs Amount", "Total Kms", "Ex. Kms", "Ex. Kms Amount", "Total"
    ]
    widths_mm = [24, 25, 34, 32, 20, 23, 28, 28, 28, 23, 23, 28, 28, 27, 34, 23, 22, 30, 25, 22, 30, 30]
    scale = doc.width / (sum(widths_mm) * mm)
    widths = [w * mm * scale for w in widths_mm]
    table_data = [[_p(h, tiny_bold) for h in headers]]
    for t in invoice.trips:
        d = t.trip_date.strftime("%d-%m-%Y") if t.trip_date else ""
        parking_toll = t.parking + t.toll + t.other_charges
        row = [
            t.ds_no, d, t.vehicle_type, t.vehicle_number, f"{t.slab_hours:g}", f"{t.slab_km:g}", f"{t.slab_rate:,.2f}",
            f"{t.extra_hour_rate:,.2f}", f"{t.extra_km_rate:,.2f}", t.start_time, t.end_time, f"{t.start_km:g}", f"{t.end_km:g}",
            f"{t.driver_bata:,.2f}", f"{parking_toll:,.2f}", f"{t.total_hours:g}", f"{t.extra_hours:g}", f"{t.extra_hour_amount:,.2f}",
            f"{t.total_km:g}", f"{t.extra_km:g}", f"{t.extra_km_amount:,.2f}", f"{t.trip_total:,.2f}"
        ]
        table_data.append([_p(v, tiny) for v in row])

    trip_table = Table(table_data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    trip_table.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 0.45, colors.black),
        ("INNERGRID", (0,0), (-1,-1), 0.2, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#eaf2ff")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("LEFTPADDING", (0,0), (-1,-1), 1.3),
        ("RIGHTPADDING", (0,0), (-1,-1), 1.3),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    story.append(trip_table)
    if any((t.notes or "").strip() for t in invoice.trips):
        notes = "; ".join([f"{t.ds_no}: {t.notes}" for t in invoice.trips if (t.notes or "").strip()])
        story.append(Spacer(1, 1 * mm))
        story.append(_p(f"Notes: {notes}", tiny))

    story.append(Spacer(1, 3 * mm))
    totals_left = [
        [_p("HSN No: 996412", tiny_bold), _p("GST NO: 36AYPPR7981L1Z8", tiny_bold)],
        [_p("Please make payment by Bank Transfer to the below account:", tiny_bold), _p("", tiny)],
        [_p("Account Name: PVR Tours & Travels<br/>SBI Account No: 39169597084<br/>IFSC: SBIN0000487", tiny), _p("As per Notification No. 22/2019 Central Tax (Rate) dated 30th September 2019, this supply is covered under REVERSE CHARGE MECHANISM. Hence, CGST / SGST payable by the recipient / receiver @ 5% on the value mentioned in the invoice.", tiny)],
    ]
    totals_right = [
        [_p("Subtotal", tiny_bold), _p(f"{invoice.subtotal:,.2f}", right)],
        [_p(f"CGST @ {invoice.cgst_rate:g}%", tiny), _p(f"{invoice.cgst:,.2f}", right)],
        [_p(f"SGST @ {invoice.sgst_rate:g}%", tiny), _p(f"{invoice.sgst:,.2f}", right)],
        [_p(f"IGST @ {invoice.igst_rate:g}%", tiny), _p(f"{invoice.igst:,.2f}", right)],
        [_p("Round Off", tiny), _p(f"{invoice.round_off:+,.2f}", right)],
        [_p("Grand Total", total_style), _p(f"{invoice.subtotal:,.2f}", total_style)],
    ]
    # meenakshi
    left_width = doc.width - 63 * mm
    bottom = Table([
        [Table(totals_left, colWidths=[65*mm, left_width - 65*mm], style=[("BOX",(0,0),(-1,-1),0.4,colors.grey),("INNERGRID",(0,0),(-1,-1),0.2,colors.lightgrey),("VALIGN",(0,0),(-1,-1),"TOP"),("SPAN",(0,1),(1,1)),("LEFTPADDING",(0,0),(-1,-1),2),("RIGHTPADDING",(0,0),(-1,-1),2),("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2)]),
         Table(totals_right, colWidths=[35*mm, 28*mm], style=[("BOX",(0,0),(-1,-1),0.4,colors.grey),("INNERGRID",(0,0),(-1,-1),0.2,colors.lightgrey),("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),2),("RIGHTPADDING",(0,0),(-1,-1),2),("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2)])]
    ], colWidths=[left_width, 63*mm])
    story.append(bottom)
    story.append(Spacer(1, 1.5 * mm))
    story.append(Table([[_p(amount_to_words_indian(invoice.grand_total), tiny_bold), _p("For PVR TOURS & TRAVELS", tiny_bold)]], colWidths=[doc.width*0.7, doc.width*0.3], style=[("BOX",(0,0),(-1,-1),0.4,colors.black),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(1,0),(1,0),"RIGHT"),("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)]))
    story.append(Spacer(1, 6 * mm))
    story.append(Table([[_p("Authorised Signatory", tiny_bold), _p("PVR TOURS & TRAVELS", tiny_bold)]], colWidths=[doc.width*0.5, doc.width*0.5], style=[("ALIGN",(0,0),(0,0),"LEFT"),("ALIGN",(1,0),(1,0),"RIGHT")]))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
def build_invoice_pdf(invoice):
    """
    Generate the final PVR Tours & Travels invoice PDF.

    PDF-only formatting changes:
    - Series and Serial No are hidden from the final PDF.
    - Invoice Number remains visible.
    - Subtotal displays the final Grand Total.
    - Separate Grand Total row is removed.
    - Font sizes are increased for better readability.
    - All 22 trip columns remain available.
    - Signature is placed on the right below "For PVR TOURS & TRAVELS".
    - Duplicate PVR TOURS & TRAVELS footer text is removed.
    - Landscape A4 is used so the complete invoice fits horizontally.
    """

    buffer = BytesIO()

    # ---------------------------------------------------------
    # A4 LANDSCAPE
    # ---------------------------------------------------------
    page = landscape(A4)

    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        rightMargin=5 * mm,
        leftMargin=5 * mm,
        topMargin=5 * mm,
        bottomMargin=5 * mm,
        title=f"PVR Invoice {invoice.invoice_number}",
        author="PVR Tours & Travels",
    )

    # ---------------------------------------------------------
    # STYLES
    # Increased font sizes for better print readability.
    # ---------------------------------------------------------
    styles = getSampleStyleSheet()

    company = ParagraphStyle(
        "company",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=19,
        leading=20,
        alignment=TA_CENTER,
        spaceAfter=1,
    )

    company_address = ParagraphStyle(
        "company_address",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
        alignment=TA_CENTER,
    )

    small_center = ParagraphStyle(
        "small_center",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=8.5,
        alignment=TA_CENTER,
    )

    small = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=8,
    )

    small_bold = ParagraphStyle(
        "small_bold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8,
    )

    tiny = ParagraphStyle(
        "tiny",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=6.4,
        leading=7.2,
    )

    tiny_bold = ParagraphStyle(
        "tiny_bold",
        parent=tiny,
        fontName="Helvetica-Bold",
        fontSize=6.5,
        leading=7.3,
    )

    section = ParagraphStyle(
        "section",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=10,
    )

    right = ParagraphStyle(
        "right",
        parent=tiny,
        alignment=TA_RIGHT,
    )

    total_style = ParagraphStyle(
        "total",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=TA_RIGHT,
    )

    signature_style = ParagraphStyle(
        "signature",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=TA_RIGHT,
    )

    # ---------------------------------------------------------
    # STORY
    # ---------------------------------------------------------
    story = []

    # ---------------------------------------------------------
    # COMPANY HEADER
    # ---------------------------------------------------------
    story.append(
        _p(
            "P.V.R. TOURS AND TRAVELS",
            company,
        )
    )

    story.append(
        _p(
            "H. No. 4-1-756, Tuljaguda, Troop Bazar, Hyderabad, "
            "Telangana State - 500 001.  "
            "Ph: 9030588882 / 9963578399",
            company_address,
        )
    )

    story.append(
        Spacer(
            1,
            2 * mm,
        )
    )

    # ---------------------------------------------------------
    # TAXABLE INVOICE
    # ---------------------------------------------------------
    story.append(
        Table(
            [
                [
                    _p(
                        "TAXABLE INVOICE",
                        section,
                    )
                ]
            ],
            colWidths=[doc.width],
            style=[
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.7,
                    colors.black,
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "LEFT",
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    2.5,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    2.5,
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    3,
                ),
            ],
        )
    )

    # ---------------------------------------------------------
    # CUSTOMER INFORMATION
    # ---------------------------------------------------------
    left = [
        ["Billed To", invoice.customer_name or ""],
        ["Address", invoice.customer_address or ""],
        ["GSTIN", invoice.customer_gstin or ""],
        ["Booked By", invoice.booked_by or ""],
        ["Used By", invoice.used_by or ""],
        ["Reference / PO", invoice.reference_number or ""],
    ]

    # IMPORTANT:
    # Series and Serial No intentionally removed from final PDF.
    right_meta = [
        [
            "Invoice Date",
            invoice.invoice_date.strftime("%d-%m-%Y")
            if invoice.invoice_date
            else "",
        ],
        [
            "Invoice No",
            invoice.invoice_number or "",
        ],
        [
            "GSTIN",
            "36AYPPR7981L1Z8",
        ],
    ]

    info_data = [
        [
            Table(
                [
                    [
                        _p(k, small_bold),
                        _p(v, small),
                    ]
                    for k, v in left
                ],
                colWidths=[
                    27 * mm,
                    115 * mm,
                ],
                style=[
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "TOP",
                    ),
                    (
                        "BOX",
                        (0, 0),
                        (-1, -1),
                        0.35,
                        colors.grey,
                    ),
                    (
                        "INNERGRID",
                        (0, 0),
                        (-1, -1),
                        0.2,
                        colors.lightgrey,
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        2,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        2,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        1.5,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        1.5,
                    ),
                ],
            ),
            Table(
                [
                    [
                        _p(k, small_bold),
                        _p(v, small),
                    ]
                    for k, v in right_meta
                ],
                colWidths=[
                    30 * mm,
                    58 * mm,
                ],
                style=[
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "TOP",
                    ),
                    (
                        "BOX",
                        (0, 0),
                        (-1, -1),
                        0.35,
                        colors.grey,
                    ),
                    (
                        "INNERGRID",
                        (0, 0),
                        (-1, -1),
                        0.2,
                        colors.lightgrey,
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        2,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        2,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        1.5,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        1.5,
                    ),
                ],
            ),
        ]
    ]

    story.append(
        Table(
            info_data,
            colWidths=[
                doc.width * 0.72,
                doc.width * 0.28,
            ],
            style=[
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    0,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    0,
                ),
            ],
        )
    )

    story.append(
        Spacer(
            1,
            2 * mm,
        )
    )

    # ---------------------------------------------------------
    # TRIP TABLE
    # ALL ORIGINAL REQUIRED FIELDS ARE RETAINED.
    # ---------------------------------------------------------
    headers = [
        "DS No",
        "Date",
        "Car Type",
        "Car No",
        "Slab Hrs",
        "Slab Kms",
        "Slab Rate",
        "Ex. Hrs Rate",
        "Ex. Kms Rate",
        "Start Time",
        "End Time",
        "Start KMS",
        "End KMS",
        "Driver Bata",
        "Parking & Toll",
        "Total Hrs",
        "Ex. Hrs",
        "Ex. Hrs Amount",
        "Total Kms",
        "Ex. Kms",
        "Ex. Kms Amount",
        "Total",
    ]

    # Original relative widths.
    # The scale automatically makes the table fit the full
    # printable A4 landscape width without horizontal overflow.
    widths_mm = [
        24,
        25,
        34,
        32,
        20,
        23,
        28,
        28,
        28,
        23,
        23,
        28,
        28,
        27,
        34,
        23,
        22,
        30,
        25,
        22,
        30,
        30,
    ]

    scale = doc.width / (
        sum(widths_mm) * mm
    )

    widths = [
        w * mm * scale
        for w in widths_mm
    ]

    table_data = [
        [
            _p(
                h,
                tiny_bold,
            )
            for h in headers
        ]
    ]

    # ---------------------------------------------------------
    # TRIP ROWS
    # ---------------------------------------------------------
    for t in invoice.trips:

        d = (
            t.trip_date.strftime("%d-%m-%Y")
            if t.trip_date
            else ""
        )

        parking_toll = (
            (t.parking or 0)
            + (t.toll or 0)
            + (t.other_charges or 0)
        )

        row = [
            t.ds_no or "",
            d,
            t.vehicle_type or "",
            t.vehicle_number or "",

            f"{(t.slab_hours or 0):g}",
            f"{(t.slab_km or 0):g}",
            f"{(t.slab_rate or 0):,.2f}",

            f"{(t.extra_hour_rate or 0):,.2f}",
            f"{(t.extra_km_rate or 0):,.2f}",

            t.start_time or "",
            t.end_time or "",

            f"{(t.start_km or 0):g}",
            f"{(t.end_km or 0):g}",

            f"{(t.driver_bata or 0):,.2f}",
            f"{parking_toll:,.2f}",

            f"{(t.total_hours or 0):g}",
            f"{(t.extra_hours or 0):g}",

            f"{(t.extra_hour_amount or 0):,.2f}",

            f"{(t.total_km or 0):g}",
            f"{(t.extra_km or 0):g}",

            f"{(t.extra_km_amount or 0):,.2f}",

            f"{(t.trip_total or 0):,.2f}",
        ]

        table_data.append(
            [
                _p(v, tiny)
                for v in row
            ]
        )

    trip_table = Table(
        table_data,
        colWidths=widths,
        repeatRows=1,
        hAlign="LEFT",
    )

    trip_table.setStyle(
        TableStyle(
            [
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.black,
                ),
                (
                    "INNERGRID",
                    (0, 0),
                    (-1, -1),
                    0.25,
                    colors.grey,
                ),
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#eaf2ff"),
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    1.2,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    1.2,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    2.5,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    2.5,
                ),
            ]
        )
    )

    story.append(trip_table)

    # ---------------------------------------------------------
    # NOTES
    # ---------------------------------------------------------
    if any(
        (t.notes or "").strip()
        for t in invoice.trips
    ):
        notes = "; ".join(
            [
                f"{t.ds_no}: {t.notes}"
                for t in invoice.trips
                if (t.notes or "").strip()
            ]
        )

        story.append(
            Spacer(
                1,
                1 * mm,
            )
        )

        story.append(
            _p(
                f"Notes: {notes}",
                tiny,
            )
        )

    story.append(
        Spacer(
            1,
            3 * mm,
        )
    )

    # ---------------------------------------------------------
    # PAYMENT / GST / TOTALS
    # ---------------------------------------------------------
    totals_left = [
        [
            _p(
                "HSN No: 996412",
                tiny_bold,
            ),
            _p(
                "GST NO: 36AYPPR7981L1Z8",
                tiny_bold,
            ),
        ],
        [
            _p(
                "Please make payment by Bank Transfer to the below account:",
                tiny_bold,
            ),
            _p(
                "",
                tiny,
            ),
        ],
        [
            _p(
                "Account Name: PVR Tours & Travels<br/>"
                "SBI Account No: 39169597084<br/>"
                "IFSC: SBIN0000487",
                tiny,
            ),
            _p(
                "As per Notification No. 22/2019 Central Tax "
                "(Rate) dated 30th September 2019, this supply "
                "is covered under REVERSE CHARGE MECHANISM. "
                "Hence, CGST / SGST payable by the recipient / "
                "receiver @ 5% on the value mentioned in the invoice.",
                tiny,
            ),
        ],
    ]

    # ---------------------------------------------------------
    # IMPORTANT:
    #
    # The actual invoice.grand_total calculation remains
    # unchanged in the database.
    #
    # ONLY the PDF DISPLAY is changed:
    #
    # Subtotal = Grand Total
    #
    # Separate Grand Total row is removed.
    # ---------------------------------------------------------
    totals_right = [
        [
            _p(
                "Subtotal",
                tiny_bold,
            ),
            _p(
                f"{invoice.subtotal:,.2f}",
                total_style,
            ),
        ],
        [
            _p(
                f"CGST @ {invoice.cgst_rate:g}%",
                tiny,
            ),
            _p(
                f"{invoice.cgst:,.2f}",
                right,
            ),
        ],
        [
            _p(
                f"SGST @ {invoice.sgst_rate:g}%",
                tiny,
            ),
            _p(
                f"{invoice.sgst:,.2f}",
                right,
            ),
        ],
        [
            _p(
                f"IGST @ {invoice.igst_rate:g}%",
                tiny,
            ),
            _p(
                f"{invoice.igst:,.2f}",
                right,
            ),
        ],
        [
            _p(
                "Round Off",
                tiny,
            ),
            _p(
                f"{invoice.round_off:+,.2f}",
                right,
            ),
        ],
         [_p("Grand Total", total_style), _p(f"{invoice.subtotal:,.2f}", total_style)],
    ]

    left_width = doc.width - 63 * mm

    bottom = Table(
        [
            [
                Table(
                    totals_left,
                    colWidths=[
                        65 * mm,
                        left_width - 65 * mm,
                    ],
                    style=[
                        (
                            "BOX",
                            (0, 0),
                            (-1, -1),
                            0.4,
                            colors.grey,
                        ),
                        (
                            "INNERGRID",
                            (0, 0),
                            (-1, -1),
                            0.2,
                            colors.lightgrey,
                        ),
                        (
                            "VALIGN",
                            (0, 0),
                            (-1, -1),
                            "TOP",
                        ),
                        (
                            "SPAN",
                            (0, 1),
                            (1, 1),
                        ),
                        (
                            "LEFTPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),
                        (
                            "RIGHTPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),
                        (
                            "TOPPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),
                        (
                            "BOTTOMPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),
                    ],
                ),
                Table(
                    totals_right,
                    colWidths=[
                        35 * mm,
                        28 * mm,
                    ],
                    style=[
                        (
                            "BOX",
                            (0, 0),
                            (-1, -1),
                            0.4,
                            colors.grey,
                        ),
                        (
                            "INNERGRID",
                            (0, 0),
                            (-1, -1),
                            0.2,
                            colors.lightgrey,
                        ),
                        (
                            "VALIGN",
                            (0, 0),
                            (-1, -1),
                            "TOP",
                        ),
                        (
                            "LEFTPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),
                        (
                            "RIGHTPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),
                        (
                            "TOPPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),
                        (
                            "BOTTOMPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),
                    ],
                ),
            ]
        ],
        colWidths=[
            left_width,
            63 * mm,
        ],
    )

    story.append(bottom)

    story.append(
        Spacer(
            1,
            2 * mm,
        )
    )

    # ---------------------------------------------------------
    # AMOUNT IN WORDS + FOR PVR TOURS & TRAVELS
    # ---------------------------------------------------------
    story.append(
        Table(
            [
                [
                    _p(
                        amount_to_words_indian(
                            invoice.grand_total
                        ),
                        small_bold,
                    ),
                    _p(
                        "For PVR TOURS & TRAVELS",
                        small_bold,
                    ),
                ]
            ],
            colWidths=[
                doc.width * 0.70,
                doc.width * 0.30,
            ],
            style=[
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.black,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (0, 0),
                    "LEFT",
                ),
                (
                    "ALIGN",
                    (1, 0),
                    (1, 0),
                    "RIGHT",
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    3,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    3,
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    3,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    3,
                ),
            ],
        )
    )

    # ---------------------------------------------------------
    # FINAL SIGNATURE
    #
    # Authorised Signatory is now on the RIGHT.
    #
    # The duplicate "PVR TOURS & TRAVELS" line below it
    # has been removed.
    # ---------------------------------------------------------
    story.append(
        Spacer(
            1,
            8 * mm,
        )
    )

    signature_table = Table(
        [
            [
                "",
                _p(
                    "Authorised Signatory",
                    signature_style,
                ),
            ]
        ],
        colWidths=[
            doc.width * 0.65,
            doc.width * 0.35,
        ],
        style=[
            (
                "ALIGN",
                (1, 0),
                (1, 0),
                "RIGHT",
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "BOTTOM",
            ),
            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                2,
            ),
            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                2,
            ),
        ],
    )

    story.append(signature_table)

    # ---------------------------------------------------------
    # BUILD PDF
    # ---------------------------------------------------------
    doc.build(story)

    buffer.seek(0)

    return buffer.getvalue()

def invoice_to_csv(invoices):
    import io
    s = io.StringIO()
    writer = csv.writer(s)
    writer.writerow([
        "Invoice Number", "Invoice Date", "Customer", "Reference", "Trips", "Subtotal", "CGST", "SGST", "IGST", "Round Off", "Grand Total"
    ])
    for inv in invoices:
        writer.writerow([
            inv.invoice_number,
            inv.invoice_date.isoformat() if inv.invoice_date else "",
            inv.customer_name,
            inv.reference_number or "",
            len(inv.trips),
            f"{inv.subtotal:.2f}",
            f"{inv.cgst:.2f}",
            f"{inv.sgst:.2f}",
            f"{inv.igst:.2f}",
            f"{inv.round_off:.2f}",
            f"{inv.grand_total:.2f}",
        ])
    return s.getvalue().encode("utf-8-sig")


def invoice_to_excel(invoices):
    """
    Export all invoices and their trip details to a production-ready Excel workbook.

    Workbook:
        1. Invoice Register
        2. Trip Details

    The export is read-only and does not modify the database.
    """

    workbook = Workbook()

    # =========================================================
    # STYLES
    # =========================================================

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    title_font = Font(
        bold=True,
        size=14
    )

    thin_side = Side(
        style="thin",
        color="D9E1F2"
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    left = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True
    )

    right = Alignment(
        horizontal="right",
        vertical="top"
    )

    # =========================================================
    # SHEET 1 - INVOICE REGISTER
    # =========================================================

    invoice_sheet = workbook.active
    invoice_sheet.title = "Invoice Register"

    invoice_headers = [
        "Database ID",
        "Invoice Number",
        "Invoice Series",
        "Invoice Serial Number",
        "Invoice Date",
        "Customer Name",
        "Customer Address",
        "Customer GSTIN",
        "Booked By",
        "Used By",
        "Reference / PO",
        "CGST %",
        "SGST %",
        "IGST %",
        "Subtotal",
        "CGST",
        "SGST",
        "IGST",
        "Round Off",
        "Grand Total",
        "Trip Count",
        "Created At",
        "Updated At",
    ]

    invoice_sheet.append(invoice_headers)

    for cell in invoice_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for inv in invoices:
        invoice_sheet.append([
            inv.id,
            inv.invoice_number or "",
            inv.invoice_series or "",
            inv.invoice_serial_number or "",
            inv.invoice_date,
            inv.customer_name or "",
            inv.customer_address or "",
            inv.customer_gstin or "",
            inv.booked_by or "",
            inv.used_by or "",
            inv.reference_number or "",
            inv.cgst_rate or 0,
            inv.sgst_rate or 0,
            inv.igst_rate or 0,
            inv.subtotal or 0,
            inv.cgst or 0,
            inv.sgst or 0,
            inv.igst or 0,
            inv.round_off or 0,
            inv.grand_total or 0,
            len(inv.trips),
            inv.created_at,
            inv.updated_at,
        ])

    # Format invoice register
    for row in invoice_sheet.iter_rows(
        min_row=2,
        max_row=invoice_sheet.max_row
    ):
        for cell in row:
            cell.border = border
            cell.alignment = left

    # Date columns
    for row in range(2, invoice_sheet.max_row + 1):
        invoice_sheet.cell(row, 5).number_format = "DD-MM-YYYY"

        invoice_sheet.cell(row, 22).number_format = "DD-MM-YYYY HH:MM:SS"
        invoice_sheet.cell(row, 23).number_format = "DD-MM-YYYY HH:MM:SS"

    # Currency columns
    currency_columns = [
        15, 16, 17, 18, 19, 20
    ]

    for row in range(2, invoice_sheet.max_row + 1):
        for col in currency_columns:
            invoice_sheet.cell(row, col).number_format = '#,##0.00'

    # Freeze header
    invoice_sheet.freeze_panes = "A2"

    # Filter
    if invoice_sheet.max_row >= 1:
        invoice_sheet.auto_filter.ref = invoice_sheet.dimensions

    # =========================================================
    # SHEET 2 - TRIP DETAILS
    # =========================================================

    trip_sheet = workbook.create_sheet("Trip Details")

    trip_headers = [
        "Trip ID",
        "Invoice ID",
        "Invoice Number",
        "DS No",
        "Trip Date",
        "Vehicle Type",
        "Vehicle Number",
        "Start Time",
        "End Time",
        "Start KM",
        "End KM",
        "Included Hours",
        "Included KM",
        "Slab Hours",
        "Slab KM",
        "Slab Rate",
        "Extra Hour Rate",
        "Extra KM Rate",
        "Extra Hours",
        "Extra KM",
        "Extra Hour Amount",
        "Extra KM Amount",
        "Base Amount",
        "Driver Bata",
        "Parking",
        "Toll",
        "Other Charges",
        "Trip Total",
        "Notes",
    ]

    trip_sheet.append(trip_headers)

    for cell in trip_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for inv in invoices:
        for trip in inv.trips:
            trip_sheet.append([
                trip.id,
                inv.id,
                inv.invoice_number or "",
                trip.ds_no or "",
                trip.trip_date,
                trip.vehicle_type or "",
                trip.vehicle_number or "",
                trip.start_time or "",
                trip.end_time or "",
                trip.start_km or 0,
                trip.end_km or 0,
                trip.total_hours or 0,
                trip.total_km or 0,
                trip.slab_hours or 0,
                trip.slab_km or 0,
                trip.slab_rate or 0,
                trip.extra_hour_rate or 0,
                trip.extra_km_rate or 0,
                trip.extra_hours or 0,
                trip.extra_km or 0,
                trip.extra_hour_amount or 0,
                trip.extra_km_amount or 0,
                trip.base_amount or 0,
                trip.driver_bata or 0,
                trip.parking or 0,
                trip.toll or 0,
                trip.other_charges or 0,
                trip.trip_total or 0,
                trip.notes or "",
            ])

    for row in trip_sheet.iter_rows(
        min_row=2,
        max_row=trip_sheet.max_row
    ):
        for cell in row:
            cell.border = border
            cell.alignment = left

    # Trip date
    for row in range(2, trip_sheet.max_row + 1):
        trip_sheet.cell(row, 5).number_format = "DD-MM-YYYY"

    # Numeric / currency columns
    currency_columns = [
        16, 17, 18, 21, 22, 23, 24, 25, 26, 27, 28
    ]

    for row in range(2, trip_sheet.max_row + 1):
        for col in currency_columns:
            trip_sheet.cell(row, col).number_format = '#,##0.00'

    trip_sheet.freeze_panes = "A2"

    if trip_sheet.max_row >= 1:
        trip_sheet.auto_filter.ref = trip_sheet.dimensions

    # =========================================================
    # COLUMN WIDTHS
    # =========================================================

    invoice_widths = {
        1: 12,
        2: 24,
        3: 18,
        4: 20,
        5: 14,
        6: 28,
        7: 35,
        8: 22,
        9: 24,
        10: 24,
        11: 24,
        12: 10,
        13: 10,
        14: 10,
        15: 15,
        16: 15,
        17: 15,
        18: 15,
        19: 15,
        20: 16,
        21: 12,
        22: 22,
        23: 22,
    }

    for col, width in invoice_widths.items():
        invoice_sheet.column_dimensions[
            get_column_letter(col)
        ].width = width

    trip_widths = {
        1: 10,
        2: 12,
        3: 24,
        4: 14,
        5: 14,
        6: 20,
        7: 20,
        8: 12,
        9: 12,
        10: 12,
        11: 12,
        12: 15,
        13: 15,
        14: 14,
        15: 14,
        16: 14,
        17: 16,
        18: 15,
        19: 14,
        20: 14,
        21: 18,
        22: 18,
        23: 15,
        24: 15,
        25: 15,
        26: 15,
        27: 18,
        28: 15,
        29: 35,
    }

    for col, width in trip_widths.items():
        trip_sheet.column_dimensions[
            get_column_letter(col)
        ].width = width

    # Header row height
    invoice_sheet.row_dimensions[1].height = 30
    trip_sheet.row_dimensions[1].height = 35

    # =========================================================
    # OUTPUT
    # =========================================================

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output.getvalue()
